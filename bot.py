import openpyxl
from webdriver_manager.chrome import ChromeDriverManager
from botcity.web import WebBot, Browser, By
from botcity.maestro import *
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
import os

BotMaestroSDK.RAISE_NOT_CONNECTED = False

# Função para converter valores de compra e venda
def converte_valores(taxa_compra, taxa_venda):
    t_compra = taxa_compra.replace(",", '.')
    t_venda = taxa_venda.replace(",", '.')
    return [float(t_compra), float(t_venda)]

# Função para acessar o site do Banco Central e coletar dados
def acessar_banco_central(bot):
    bot.browse("https://www.bcb.gov.br/estabilidadefinanceira/fechamentodolar")
    
    # Localiza o iframe correto
    iframe = bot.find_element(
        "/html/body/app-root/app-root/div/div/main/dynamic-comp/div/div[1]/iframe", By.XPATH)
    bot.enter_iframe(iframe)

    # Espera os dados aparecerem na página
    while len(bot.find_elements('/html/body/div[2]/table/tbody/tr[2]', By.XPATH)) < 1:
        bot.wait(1000)
        print('Carregando...')

    # Coleta os dados de compra e venda
    data = bot.find_element(
        '/html/body/div[2]/table/tbody/tr[2]/td[1]', By.XPATH).text
    taxa_compra = bot.find_element(
        '/html/body/div[2]/table/tbody/tr[2]/td[2]', By.XPATH).text
    taxa_venda = bot.find_element(
        '/html/body/div[2]/table/tbody/tr[2]/td[3]', By.XPATH).text

    valores_convertidos = converte_valores(taxa_compra, taxa_venda)
    return [data, valores_convertidos[0], valores_convertidos[1]]

# Função para salvar os dados na planilha Excel
def salvar_planilha(dados):
    arquivo = 'dados.xlsx'

    if os.path.exists(arquivo):
        workbook = load_workbook(arquivo)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Adiciona o cabeçalho na primeira linha
        sheet.append(["Data", "Compra", "Venda"])

    # Adiciona os novos dados (Data, Compra, Venda)
    sheet.append([dados[0], dados[1], dados[2]])

    workbook.save(arquivo)
    return workbook

# Função para gerar o gráfico de linha na planilha
def gerar_grafico(planilha):
    sheet = planilha.active
    chart = LineChart()#Cria um objeto de gráfico do tipo LineChart
    chart.title = "Taxa de Compra e Venda do Dólar"
    chart.style = 20
    chart.y_axis.title = 'Taxa R$'
    chart.x_axis.title = 'Data'
    
    # Definindo os dados para o gráfico (colunas de Compra e Venda)
    
    data = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=sheet.max_row)
    
    # Definindo as categorias (coluna de Data)
    categorias = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    
    # Adicionando os dados ao gráfico
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categorias)
    
     # Verifica se já existe um gráfico na posição especificada, para não sobrepor
    if "E5" not in [x.anchor for x in sheet._charts]:
        sheet.add_chart(chart, "E5")
    else:
        sheet._charts.clear()  # Remove o gráfico anterior
        sheet.add_chart(chart, "E5")

    planilha.save('dados.xlsx')

# Função principal
def main():
    maestro = BotMaestroSDK.from_sys_args()
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    bot = WebBot()
    bot.headless = False

    bot.browser = Browser.CHROME
    bot.driver_path = ChromeDriverManager().install()

    try:
        # Acessa o site e coleta os dados
        dados = acessar_banco_central(bot)
        # Salva os dados na planilha
        workbook = salvar_planilha(dados)
        # Gera o gráfico (ou atualiza) na planilha
        gerar_grafico(workbook)

        print("Dados coletados e gráfico gerado com sucesso.")

    except Exception as ex:
        print('Falhou')
        bot.save_screenshot('error.png')
        print(ex)

    bot.stop_browser()

if __name__ == "__main__":
    main()